"""발주 엑셀 양식 만들기. GUI와 CLI가 같이 쓴다.

exe만 받은 사장님은 양식을 구할 방법이 없다 — 저장소도 없고 CLI도 못 쓴다.
그래서 앱 안에서 바로 만들 수 있어야 한다.
"""

from __future__ import annotations

from pathlib import Path

from excel_reader import COLUMNS

DEFAULT_NAME = "발주_양식.xlsx"

# 두 번째 줄에 들어갈 예시. 지우고 쓰라고 안내한다.
EXAMPLE = [
    "https://wellrootb2b.com/product/detail.html?product_no=675",
    "(없음)",
    2,
    "홍길동",
    "010-0000-0000",  # 전화번호(일반전화) — 필수. 휴대폰 번호를 적어도 된다
    "",  # 휴대폰번호 — 선택. 비우면 주문서의 휴대전화도 빈칸이 된다
    "12345",
    "서울특별시 강남구 테헤란로 152",
    "3층",
    "문 앞",
]

REQUIRED_COLUMNS = ("상품url", "수량", "수취인", "전화번호", "우편번호", "주소")


def write_template(path: str | Path) -> Path:
    """양식을 만들어 저장한 경로를 돌려준다."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    path = Path(path)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "발주"

    sheet.append(list(COLUMNS))
    sheet.append(EXAMPLE)

    # 필수 칸을 눈에 띄게 — 사장님이 무엇을 꼭 채워야 하는지 한눈에 보이게 한다
    required = PatternFill("solid", fgColor="FFF2CC")
    for i, name in enumerate(COLUMNS, start=1):
        cell = sheet.cell(row=1, column=i)
        cell.font = Font(bold=True)
        if name in REQUIRED_COLUMNS:
            cell.fill = required
        sheet.column_dimensions[cell.column_letter].width = max(14, len(name) + 8)

    sheet.freeze_panes = "A2"
    workbook.save(path)
    return path


HELP_TEXT = (
    "발주 양식을 만들었습니다.\n\n"
    "· 색이 칠해진 칸은 반드시 채워야 합니다\n"
    "   (상품url · 수량 · 수취인 · 전화번호 · 우편번호 · 주소)\n"
    "· 휴대폰번호 · 상세주소 · 배송메모는 비워도 됩니다\n"
    "· 옵션이 없는 상품은 (없음) 으로 두세요\n"
    "· 수취인과 주소가 같은 줄끼리 한 주문으로 묶입니다\n\n"
    "2번째 줄은 예시입니다. 지우고 실제 발주를 채워주세요."
)
