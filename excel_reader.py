"""발주 엑셀을 읽어 검증하고 배송지별로 묶는다.

이 모듈은 사이트 DOM에 의존하지 않으므로 몰 스킨이 바뀌어도 영향받지 않는다.
브라우저를 띄우기 전에 여기서 최대한 걸러낸다 — 반쯤 진행된 뒤 실패하는 게 가장 나쁘다.
"""

from __future__ import annotations

import re
from dataclasses import dataclass, field
from pathlib import Path

from openpyxl import load_workbook

# 표준 헤더 → 내부 필드명. 템플릿도 이 순서로 만든다.
# 헤더는 공백/대소문자를 무시하고 매칭한다.
COLUMNS = {
    "상품url": "product_url",
    "옵션": "option",
    "수량": "quantity",
    "수취인": "receiver",
    "전화번호": "tel",  # 주문서의 '일반전화'
    "휴대폰번호": "phone",  # 주문서의 '휴대전화'
    "우편번호": "zipcode",
    "주소": "address",
    "상세주소": "address_detail",
    "배송메모": "memo",
}

# 예전 양식·다른 표기도 받아준다. `연락처`는 휴대폰번호 하나만 쓰던 초기 양식.
ALIASES = {
    "연락처": "phone",
    "휴대폰": "phone",
    "휴대전화": "phone",
    "핸드폰번호": "phone",
    "전화": "tel",
    "일반전화": "tel",
    "유선전화": "tel",
}

# 필수 컬럼. 주문서(카페24)의 검증 규칙과 맞춰 놓은 것이다 —
# 수취자 전화번호는 `isFill`(필수), 핸드폰번호는 `isNumber`(선택)이다.
REQUIRED = ("product_url", "quantity", "receiver", "tel", "zipcode", "address")

# 빈칸이어도 되는 컬럼. 비어 있으면 주문서의 해당 칸도 **비운다**(아래 주의 참고).
OPTIONAL = ("phone", "address_detail", "memo", "option")

# 옵션 칸에 "없음"류가 적혀 있으면 옵션 미선택으로 본다.
# 실제 발주 엑셀에서 `(없음)`으로 쓰고 있어서 그대로 받아준다.
NO_OPTION = {"", "(없음)", "없음", "없슴", "-", "無", "x", "X"}

# 주문서의 '휴대전화' 앞자리 select에 실제로 들어있는 값(실측).
# 전화번호/휴대폰번호 칸을 바꿔 넣는 실수를 브라우저 띄우기 전에 잡는 데 쓴다.
MOBILE_PREFIXES = ("010", "011", "016", "017", "018", "019",
                   "0502", "0503", "0504", "0505", "0506", "0507", "0508")

# 카페24 상품 URL에서 상품번호를 뽑는다.
#   /product/상품명/1234/category/56/display/1/  →  1234
#   ...?product_no=1234                          →  1234
_PRODUCT_NO_PATTERNS = (
    re.compile(r"/product/[^/]+/(\d+)"),
    re.compile(r"[?&]product_no=(\d+)"),
)


@dataclass
class OrderLine:
    """엑셀 한 줄 = 상품 하나."""

    row: int  # 엑셀 행 번호 (1-based). 오류 리포트에 그대로 쓴다.
    # --- 필수 ---
    product_url: str
    quantity: int
    receiver: str
    tel: str  # 전화번호 → 주문서 '일반전화' (몰이 필수로 요구)
    zipcode: str
    address: str
    # --- 선택 (빈칸이면 주문서의 해당 칸도 비운다) ---
    phone: str = ""  # 휴대폰번호 → 주문서 '휴대전화'
    option: str = ""
    address_detail: str = ""
    memo: str = ""

    @property
    def product_no(self) -> str | None:
        for pattern in _PRODUCT_NO_PATTERNS:
            m = pattern.search(self.product_url)
            if m:
                return m.group(1)
        return None

    @property
    def shipping_key(self) -> tuple[str, str, str, str]:
        """배송지 동일성 판단 기준. 같으면 한 주문으로 묶는다.

        전화번호도 포함한다 — 빼면 값이 다른 줄끼리 묶이면서 한쪽 번호가 조용히 사라진다.
        묶이길 원하면 두 칸을 같게 적으면 되고, 어긋나면 --dry-run에서 눈에 띈다.
        """
        return (
            _squash(self.receiver),
            _digits(self.phone),
            _digits(self.tel),
            _squash(self.address + self.address_detail),
        )


@dataclass
class OrderGroup:
    """같은 배송지로 가는 상품들 = 주문 1건."""

    receiver: str
    tel: str
    zipcode: str
    address: str
    phone: str = ""
    address_detail: str = ""
    memo: str = ""
    lines: list[OrderLine] = field(default_factory=list)


class ExcelError(Exception):
    """사용자에게 그대로 보여줄 엑셀 오류. 메시지에 행 번호를 담는다."""


def read_orders(path: str | Path) -> list[OrderGroup]:
    """엑셀을 읽어 배송지별 주문 묶음으로 반환한다. 형식 오류면 ExcelError."""
    lines = _read_lines(path)
    if not lines:
        raise ExcelError("주문 라인이 하나도 없습니다. 헤더 아래에 데이터를 넣어주세요.")
    return _group_by_shipping(lines)


def _read_lines(path: str | Path) -> list[OrderLine]:
    # read_only=True로 열면 대용량에서 빠르지만, 여기서는 10건 미만이라 이점이 없고
    # data_only=True로 수식 대신 계산된 값을 읽는 게 훨씬 중요하다.
    workbook = load_workbook(filename=path, data_only=True)
    sheet = workbook.active

    rows = sheet.iter_rows(values_only=True)
    try:
        header = next(rows)
    except StopIteration:
        raise ExcelError("빈 파일입니다.") from None

    index = _map_header(header)
    missing = [k for k in REQUIRED if k not in index]
    if missing:
        names = ", ".join(_korean_name(k) for k in missing)
        raise ExcelError(f"필수 컬럼이 없습니다: {names}")

    lines: list[OrderLine] = []
    errors: list[str] = []

    for offset, row in enumerate(rows, start=2):  # 헤더가 1행
        if all(cell is None or str(cell).strip() == "" for cell in row):
            continue  # 빈 줄은 조용히 건너뛴다

        values = {key: _cell(row, pos) for key, pos in index.items()}

        blank = [_korean_name(k) for k in REQUIRED if not values.get(k)]
        if blank:
            names = ", ".join(blank)
            errors.append(f"{offset}행: {names}{_subject_particle(blank[-1])} 비어 있습니다")
            continue

        quantity = _parse_quantity(values["quantity"])
        if quantity is None:
            errors.append(f"{offset}행: 수량 '{values['quantity']}'을(를) 숫자로 읽을 수 없습니다")
            continue

        url = values["product_url"]
        if not url.startswith("http"):
            errors.append(f"{offset}행: 상품URL이 http로 시작하지 않습니다 — '{url}'")
            continue

        # 휴대폰번호는 선택이지만, 적었다면 휴대전화 번호여야 한다.
        # (주문서의 휴대전화 앞자리 select에 지역번호가 아예 없어서 넣을 수가 없다)
        phone = values.get("phone", "")
        if phone and not _is_mobile(phone):
            errors.append(
                f"{offset}행: 휴대폰번호 '{phone}'는 휴대전화 번호가 아닙니다. "
                "전화번호 칸과 바뀌지 않았는지 확인하세요"
            )
            continue

        lines.append(
            OrderLine(
                row=offset,
                product_url=url,
                quantity=quantity,
                receiver=values["receiver"],
                tel=values["tel"],
                address=values["address"],
                zipcode=values["zipcode"],
                phone=phone,
                option=_normalize_option(values.get("option", "")),
                address_detail=values.get("address_detail", ""),
                memo=values.get("memo", ""),
            )
        )

    if errors:
        raise ExcelError("엑셀에 문제가 있습니다:\n  - " + "\n  - ".join(errors))

    return lines


def _group_by_shipping(lines: list[OrderLine]) -> list[OrderGroup]:
    groups: dict[tuple[str, str, str], OrderGroup] = {}
    for line in lines:
        key = line.shipping_key
        group = groups.get(key)
        if group is None:
            group = OrderGroup(
                receiver=line.receiver,
                phone=line.phone,
                tel=line.tel,
                zipcode=line.zipcode,
                address=line.address,
                address_detail=line.address_detail,
                memo=line.memo,
            )
            groups[key] = group
        group.lines.append(line)
    return list(groups.values())


def _map_header(header: tuple) -> dict[str, int]:
    index: dict[str, int] = {}
    for pos, cell in enumerate(header):
        if cell is None:
            continue
        name = _squash(str(cell)).lower()
        key = COLUMNS.get(name) or ALIASES.get(name)
        if key and key not in index:
            index[key] = pos
    return index


def _is_mobile(phone: str) -> bool:
    """휴대폰번호 칸에 일반전화(02-, 031- 등)가 들어간 걸 잡는다."""
    digits = _digits(phone)
    return digits.startswith(MOBILE_PREFIXES)


def _cell(row: tuple, pos: int) -> str:
    if pos >= len(row) or row[pos] is None:
        return ""
    return str(row[pos]).strip()


def _parse_quantity(raw: str) -> int | None:
    try:
        # 엑셀이 숫자를 "3.0"으로 주는 경우가 흔하다
        quantity = int(float(raw))
    except (TypeError, ValueError):
        return None
    return quantity if quantity > 0 else None


def _subject_particle(word: str) -> str:
    """받침이 있으면 '이', 없으면 '가'. 오류 메시지를 자연스럽게 읽히게 한다."""
    if not word:
        return "가"
    last = word[-1]
    if "가" <= last <= "힣":
        return "이" if (ord(last) - 0xAC00) % 28 else "가"
    return "가"


def _normalize_option(raw: str) -> str:
    return "" if raw.strip() in NO_OPTION else raw.strip()


def _korean_name(key: str) -> str:
    for korean, internal in COLUMNS.items():
        if internal == key:
            return korean
    return key


def _squash(text: str) -> str:
    return re.sub(r"\s+", "", text or "")


def _digits(text: str) -> str:
    return re.sub(r"\D", "", text or "")
